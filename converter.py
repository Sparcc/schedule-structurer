import re
from openpyxl import load_workbook
import plotly.plotly as py
import plotly.graph_objs as go
import plotly

username = 'sparcccode'
api_key = 'tKQTWqVGh3ilSRXXqJEr'
plotly.tools.set_credentials_file(username=username, api_key=api_key)

wb = load_workbook("schedule.xlsx")
ws = wb.active
startingCells = 4
maxCells = 261
numData = maxCells-startingCells
regexCode = '(\d{1,3}) (\d{0,3}%) (.+) (\d{0,3} days{0,1}|\d{0,1}.\d{0,1}.\d{0,1} days{0,1}) (\S{3} \d{1,2}\/\d{1,2}\/\d{1,2}) (\S{3} \d{1,2}\/\d{1,2}\/\d{1,2})( \d{1,3}|)(.*){0,1}'
class ScheduleEntry:
    ID =""
    Completion = ""
    Task = ""
    Duration = ""
    Start = ""
    Finish = ""
    Predecessors = ""
    Resources = "Not Defined"

numMismatches = 0
schedule = []
i=startingCells
while i < maxCells:
    text = ws["A"+str(i)].value
    #print(text)
    result = re.search(regexCode, text)
    if result:
        #print("Match Found")
        resultGroups = result.groups()
        #print(resultGroups)
        if len(resultGroups) > 2:
            scheduleEntry = ScheduleEntry()
            scheduleEntry.ID = str(resultGroups[0])
            scheduleEntry.Completion = str(resultGroups[1])
            scheduleEntry.Task = str(resultGroups[2])
            scheduleEntry.Duration = str(resultGroups[3])
            scheduleEntry.Start = str(resultGroups[4])
            scheduleEntry.Finish = str(resultGroups[5])
            
            if len(resultGroups) > 5:
                scheduleEntry.Predecessors = resultGroups[6]
                resultGroups[6]
                if len(resultGroups) > 6:
                    scheduleEntry.Resources = resultGroups[7]
                    resultGroups[7]
            schedule.append(scheduleEntry)
            
    else:
        numMismatches +=1
    i+=1

if numMismatches >0:
    print("Could not match {numMismatches}/{numData} rows"
    .format(numMismatches=numMismatches,numData=numData))
    
tasksComplete = 0
tasksInProgress = 0
tasksNotStarted = 0


for value in schedule:
    intValue = int(value.Completion[:len(value.Completion)-1])
    if intValue == 100:
        tasksComplete +=1
    elif intValue > 0 and intValue <100:
        tasksInProgress +=1
    elif intValue == 0:
        tasksNotStarted +=1

labels = ['tasksComplete','tasksInProgress','tasksNotStarted']
values = [tasksComplete,tasksInProgress,tasksNotStarted]

trace = go.Pie(labels=labels, values=values)
data = [trace]
layout = go.Layout(title='Summary of Tasks', width=800, height=640)
fig = go.Figure(data=data, layout=layout)
#py.iplot([trace], filename='scheduleOverview')
py.image.save_as(fig, filename='task-summary.png')

'''
for value in schedule:
    searchForMe = '(.*Thomas Rea.*)'
    result = re.search(searchForMe,value.Resources)
    if result:
        print(#value.ID+", "+
        value.Completion+", "+
        value.Task+", "+
        value.Duration+", "+
        value.Start+", "+
        value.Finish+", "+
        #value.Predecessors+", "+
        value.Resources)
'''